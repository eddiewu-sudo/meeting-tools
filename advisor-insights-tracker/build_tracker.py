#!/usr/bin/env python3
"""
Build the styled advisor-insights tracker .xlsx from extracted rows.

Input: a JSON file containing a list of rows. Each row is a list of 6 strings
in this exact order:
    [Session, 日期, Key Insight (1句話), Detail / Verbatim, Action Item, Status]

Usage:
    python build_tracker.py --rows rows.json --out /mnt/user-data/outputs/Advisor_Insights.xlsx \
        [--title "Advisor Insights"]

The output keeps the same six columns, header styling, per-session colour tint,
wrapped text, column widths, and frozen header row used throughout, so every
weekly run looks identical and pastes cleanly into Google Sheets.
"""
import argparse
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = ["Session", "日期", "Key Insight (1句話)", "Detail / Verbatim", "Action Item", "Status"]

# Stable palette so repeated sessions keep a consistent tint across weeks.
SESSION_FILLS = [
    "E8F0FE",  # blue
    "FCE8E6",  # red
    "E6F4EA",  # green
    "FEF7E0",  # amber
    "F3E8FD",  # purple
    "E0F7FA",  # cyan
]


def session_color(session, mapping):
    if session not in mapping:
        mapping[session] = SESSION_FILLS[len(mapping) % len(SESSION_FILLS)]
    return mapping[session]


def safe_sheet_title(title):
    title = title or "Advisor Insights"
    for ch in r':\/?*[]':
        title = title.replace(ch, "-")
    return title[:31]


def build(rows, out_path, title):
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(title)

    ws.append(HEADERS)
    for r in rows:
        if len(r) != 6:
            raise ValueError(f"Each row must have exactly 6 cells; got {len(r)}: {r}")
        ws.append(r)

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    mapping = {}
    for ridx in range(2, ws.max_row + 1):
        session = ws.cell(row=ridx, column=1).value
        for c in range(1, 7):
            cell = ws.cell(row=ridx, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = center if c in (1, 2, 6) else left_wrap
        ws.cell(row=ridx, column=1).fill = PatternFill(
            "solid", start_color=session_color(session, mapping)
        )

    for col, w in {"A": 9, "B": 12, "C": 34, "D": 60, "E": 40, "F": 22}.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    wb.save(out_path)
    return ws.max_row - 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", required=True, help="Path to JSON file: list of 6-cell rows")
    ap.add_argument("--out", required=True, help="Output .xlsx path")
    ap.add_argument("--title", default="Advisor Insights", help="Worksheet/title name")
    args = ap.parse_args()

    with open(args.rows, encoding="utf-8") as f:
        rows = json.load(f)

    if not isinstance(rows, list) or not rows:
        print("No rows to write. Did the date window match any session?", file=sys.stderr)
        sys.exit(2)

    n = build(rows, args.out, args.title)
    print(json.dumps({"status": "ok", "rows_written": n, "out": args.out}))


if __name__ == "__main__":
    main()
